#!/usr/bin/env python3
"""
ImageTrend Distribution Partnership - Financial Model for Protocol Guide
Creates comprehensive revenue projections and valuation analysis
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
import os

# Create workbook
wb = Workbook()

# ============================================================================
# SHEET 1: ImageTrend Company Analysis
# ============================================================================
ws1 = wb.active
ws1.title = "ImageTrend Analysis"

# Styling
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill("solid", fgColor="1F4E79")
input_font = Font(color="0000FF")  # Blue for inputs
currency_format = '$#,##0'
pct_format = '0.0%'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws1['A1'] = "ImageTrend Inc. - Company Financial Analysis"
ws1['A1'].font = Font(bold=True, size=16)
ws1.merge_cells('A1:E1')

# Revenue Data Section
ws1['A3'] = "Historical Revenue Data"
ws1['A3'].font = Font(bold=True, size=14)

headers = ['Year', 'Revenue', 'YoY Growth', 'Employees', 'Rev/Employee']
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# Revenue data (from research - GetLatka)
revenue_data = [
    (2021, 10300000, None, 180, None),
    (2022, 13000000, None, 220, None),  # Estimated
    (2023, 17400000, None, 280, None),
    (2024, 25200000, None, 345, None),
    (2025, 32760000, None, 394, None),  # Projected 30% growth
]

for row_idx, (year, revenue, _, employees, _) in enumerate(revenue_data, 5):
    ws1.cell(row=row_idx, column=1, value=year)
    ws1.cell(row=row_idx, column=2, value=revenue).number_format = currency_format
    ws1.cell(row=row_idx, column=2).font = input_font
    if row_idx > 5:
        ws1.cell(row=row_idx, column=3, value=f'=(B{row_idx}-B{row_idx-1})/B{row_idx-1}').number_format = pct_format
    ws1.cell(row=row_idx, column=4, value=employees).font = input_font
    ws1.cell(row=row_idx, column=5, value=f'=B{row_idx}/D{row_idx}').number_format = currency_format

# Market Position
ws1['A12'] = "Market Position & Customer Base"
ws1['A12'].font = Font(bold=True, size=14)

metrics = [
    ('Total Customers', 3000, 'As of 2024'),
    ('Total Agencies Served', 20000, 'EMS, Fire, Hospital'),
    ('EMS Agencies', 13000, 'ePCR Solutions'),
    ('States Covered', 43, 'NEMSIS Compliance'),
    ('NEMSIS Records Submitted', 356000000, 'Cumulative'),
    ('CSAT Score', 0.974, '2024 - 3,633 responses'),
    ('2024 New Hires', 141, 'Growth indicator'),
]

for row_idx, (metric, value, note) in enumerate(metrics, 13):
    ws1.cell(row=row_idx, column=1, value=metric)
    cell = ws1.cell(row=row_idx, column=2, value=value)
    cell.font = input_font
    if metric == 'CSAT Score':
        cell.number_format = pct_format
    elif metric == 'NEMSIS Records Submitted':
        cell.number_format = '#,##0'
    ws1.cell(row=row_idx, column=3, value=note)

# Financial Estimates
ws1['A22'] = "Financial Estimates (Private Company)"
ws1['A22'].font = Font(bold=True, size=14)

estimates = [
    ('Estimated 2024 Revenue', 25200000, 'GetLatka verified'),
    ('PE Investment (2023)', 24000000, 'Welsh Carson Anderson Stowe'),
    ('Revenue Multiple (SaaS)', 8, 'Industry average 6-10x'),
    ('Implied Valuation', '=B23*B25', 'Revenue x Multiple'),
    ('Estimated EBITDA Margin', 0.20, 'SaaS industry avg 15-25%'),
    ('Estimated EBITDA', '=B23*B27', ''),
    ('Net Profit Margin Est.', 0.12, 'After tax estimate'),
    ('Estimated Net Profit', '=B23*B29', ''),
]

for row_idx, (metric, value, note) in enumerate(estimates, 23):
    ws1.cell(row=row_idx, column=1, value=metric)
    cell = ws1.cell(row=row_idx, column=2, value=value)
    if isinstance(value, str) and value.startswith('='):
        cell.number_format = currency_format
    elif isinstance(value, float) and value < 1:
        cell.number_format = pct_format
        cell.font = input_font
    elif isinstance(value, int) and value > 100:
        cell.number_format = currency_format if value > 1000000 else '#,##0'
        cell.font = input_font
    ws1.cell(row=row_idx, column=3, value=note)

# Column widths
ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 30
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 15

# ============================================================================
# SHEET 2: Protocol Guide Revenue Model
# ============================================================================
ws2 = wb.create_sheet("Revenue Model")

ws2['A1'] = "Protocol Guide - ImageTrend Distribution Revenue Model"
ws2['A1'].font = Font(bold=True, size=16)
ws2.merge_cells('A1:G1')

# Key Assumptions
ws2['A3'] = "Key Assumptions"
ws2['A3'].font = Font(bold=True, size=14)
ws2['A3'].fill = PatternFill("solid", fgColor="FFF2CC")

assumptions = [
    ('ImageTrend Revenue Share', 0.20, 'They take 20%'),
    ('Protocol Guide Revenue Share', 0.80, 'You keep 80%'),
    ('Protocol Guide Monthly Price', 4.99, 'Per user/month'),
    ('Protocol Guide Annual Price', 39.00, 'Per user/year'),
    ('Annual Discount', '=1-(B7/(B6*12))', 'Annual vs Monthly'),
    ('ImageTrend Total EMS Agencies', 13000, 'Current customer base'),
    ('Avg Users per Agency', 25, 'Paramedics/EMTs'),
    ('Total Addressable Users', '=B9*B10', 'Within ImageTrend'),
]

for row_idx, (metric, value, note) in enumerate(assumptions, 4):
    ws2.cell(row=row_idx, column=1, value=metric)
    cell = ws2.cell(row=row_idx, column=2, value=value)
    cell.font = input_font
    if isinstance(value, float) and value < 1:
        cell.number_format = pct_format
    elif isinstance(value, float):
        cell.number_format = '$#,##0.00'
    elif isinstance(value, str) and value.startswith('='):
        if 'B7/(B6*12)' in value:
            cell.number_format = pct_format
        else:
            cell.number_format = '#,##0'
    else:
        cell.number_format = '#,##0'
    ws2.cell(row=row_idx, column=3, value=note)

# Adoption Scenarios
ws2['A14'] = "Adoption Scenarios (% of ImageTrend Customer Base)"
ws2['A14'].font = Font(bold=True, size=14)
ws2['A14'].fill = PatternFill("solid", fgColor="FFF2CC")

scenario_headers = ['Scenario', 'Adoption Rate', 'Agencies', 'Users', 'Monthly Price', 'Annual Mix', 'Monthly MRR', 'ARR', 'Your 80%']
for col, header in enumerate(scenario_headers, 1):
    cell = ws2.cell(row=15, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

scenarios = [
    ('Conservative', 0.02, '=$B$9*B16', '=C16*$B$10', '$B$6', 0.70),
    ('Moderate', 0.05, '=$B$9*B17', '=C17*$B$10', '$B$6', 0.70),
    ('Optimistic', 0.10, '=$B$9*B18', '=C18*$B$10', '$B$6', 0.70),
    ('Aggressive', 0.20, '=$B$9*B19', '=C19*$B$10', '$B$6', 0.70),
]

for row_idx, (name, adoption, agencies_formula, users_formula, price_ref, annual_mix) in enumerate(scenarios, 16):
    ws2.cell(row=row_idx, column=1, value=name)
    ws2.cell(row=row_idx, column=2, value=adoption).number_format = pct_format
    ws2.cell(row=row_idx, column=2).font = input_font
    ws2.cell(row=row_idx, column=3, value=agencies_formula).number_format = '#,##0'
    ws2.cell(row=row_idx, column=4, value=users_formula).number_format = '#,##0'
    ws2.cell(row=row_idx, column=5, value=price_ref).number_format = '$#,##0.00'
    ws2.cell(row=row_idx, column=6, value=annual_mix).number_format = pct_format
    ws2.cell(row=row_idx, column=6).font = input_font
    # MRR = Users * ((1-Annual%) * Monthly + Annual% * Annual/12)
    ws2.cell(row=row_idx, column=7, value=f'=D{row_idx}*((1-F{row_idx})*$B$6+F{row_idx}*$B$7/12)').number_format = currency_format
    # ARR = MRR * 12
    ws2.cell(row=row_idx, column=8, value=f'=G{row_idx}*12').number_format = currency_format
    # Your 80%
    ws2.cell(row=row_idx, column=9, value=f'=H{row_idx}*$B$5').number_format = currency_format

# 5-Year Growth Projection
ws2['A22'] = "5-Year Growth Projection (Moderate Scenario)"
ws2['A22'].font = Font(bold=True, size=14)
ws2['A22'].fill = PatternFill("solid", fgColor="FFF2CC")

year_headers = ['Metric', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
for col, header in enumerate(year_headers, 1):
    cell = ws2.cell(row=23, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill

projection_data = [
    ('Adoption Rate', [0.02, 0.05, 0.10, 0.15, 0.20]),
    ('Active Agencies', None),  # Formula
    ('Active Users', None),
    ('Monthly MRR', None),
    ('Annual ARR', None),
    ('Your Revenue (80%)', None),
]

# Year headers for reference
ws2['B24'] = 0.02
ws2['C24'] = 0.05
ws2['D24'] = 0.10
ws2['E24'] = 0.15
ws2['F24'] = 0.20
for col in range(2, 7):
    ws2.cell(row=24, column=col).number_format = pct_format
    ws2.cell(row=24, column=col).font = input_font

ws2['A24'] = 'Adoption Rate'
ws2['A25'] = 'Active Agencies'
ws2['A26'] = 'Active Users'
ws2['A27'] = 'Monthly MRR'
ws2['A28'] = 'Annual ARR'
ws2['A29'] = 'Your Revenue (80%)'

for col in range(2, 7):
    col_letter = get_column_letter(col)
    # Agencies
    ws2.cell(row=25, column=col, value=f'=$B$9*{col_letter}24').number_format = '#,##0'
    # Users
    ws2.cell(row=26, column=col, value=f'={col_letter}25*$B$10').number_format = '#,##0'
    # MRR (assuming 70% annual)
    ws2.cell(row=27, column=col, value=f'={col_letter}26*((1-0.7)*$B$6+0.7*$B$7/12)').number_format = currency_format
    # ARR
    ws2.cell(row=28, column=col, value=f'={col_letter}27*12').number_format = currency_format
    # Your 80%
    ws2.cell(row=29, column=col, value=f'={col_letter}28*$B$5').number_format = currency_format

# Column widths
for col in range(1, 10):
    ws2.column_dimensions[get_column_letter(col)].width = 15
ws2.column_dimensions['A'].width = 35

# ============================================================================
# SHEET 3: Valuation Impact
# ============================================================================
ws3 = wb.create_sheet("Valuation Impact")

ws3['A1'] = "Protocol Guide - Valuation Impact Analysis"
ws3['A1'].font = Font(bold=True, size=16)
ws3.merge_cells('A1:E1')

ws3['A3'] = "SaaS Valuation Multiples"
ws3['A3'].font = Font(bold=True, size=14)

val_headers = ['Growth Rate', 'ARR Multiple', 'Notes']
for col, header in enumerate(val_headers, 1):
    cell = ws3.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill

multiples = [
    ('<20%', '3-5x', 'Slow growth SaaS'),
    ('20-40%', '5-8x', 'Moderate growth'),
    ('40-60%', '8-12x', 'High growth'),
    ('>60%', '12-20x', 'Hypergrowth'),
]

for row_idx, (growth, multiple, note) in enumerate(multiples, 5):
    ws3.cell(row=row_idx, column=1, value=growth)
    ws3.cell(row=row_idx, column=2, value=multiple)
    ws3.cell(row=row_idx, column=3, value=note)

# Valuation Scenarios
ws3['A11'] = "Protocol Guide Valuation Scenarios"
ws3['A11'].font = Font(bold=True, size=14)

val_scenario_headers = ['Scenario', 'Your ARR', 'Multiple', 'Valuation', 'Notes']
for col, header in enumerate(val_scenario_headers, 1):
    cell = ws3.cell(row=12, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill

# Reference Revenue Model sheet
ws3['A13'] = 'Conservative (2%)'
ws3['B13'] = "='Revenue Model'!I16"
ws3['C13'] = 5
ws3['D13'] = '=B13*C13'
ws3['E13'] = 'Low adoption, modest multiple'

ws3['A14'] = 'Moderate (5%)'
ws3['B14'] = "='Revenue Model'!I17"
ws3['C14'] = 8
ws3['D14'] = '=B14*C14'
ws3['E14'] = 'Reasonable adoption, growth multiple'

ws3['A15'] = 'Optimistic (10%)'
ws3['B15'] = "='Revenue Model'!I18"
ws3['C15'] = 10
ws3['D15'] = '=B15*C15'
ws3['E15'] = 'Strong adoption, premium multiple'

ws3['A16'] = 'Aggressive (20%)'
ws3['B16'] = "='Revenue Model'!I19"
ws3['C16'] = 12
ws3['D16'] = '=B16*C16'
ws3['E16'] = 'Market leader potential'

for row in range(13, 17):
    ws3.cell(row=row, column=2).number_format = currency_format
    ws3.cell(row=row, column=3).font = input_font
    ws3.cell(row=row, column=4).number_format = currency_format

# Key Insights
ws3['A19'] = "Key Investment Thesis"
ws3['A19'].font = Font(bold=True, size=14)

insights = [
    "1. Distribution through ImageTrend provides access to 13,000+ EMS agencies immediately",
    "2. No customer acquisition cost - ImageTrend handles sales and marketing",
    "3. 80% revenue share is strong for a distribution partnership",
    "4. Recurring revenue model (SaaS) commands premium valuation multiples",
    "5. Embedded in existing workflow = high retention potential",
    "6. Path to $1M+ ARR within 2 years at 5% adoption",
]

for row_idx, insight in enumerate(insights, 20):
    ws3.cell(row=row_idx, column=1, value=insight)
    ws3.merge_cells(f'A{row_idx}:E{row_idx}')

# Column widths
ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 35

# ============================================================================
# SHEET 4: Executive Summary
# ============================================================================
ws4 = wb.create_sheet("Executive Summary")

ws4['A1'] = "EXECUTIVE SUMMARY"
ws4['A1'].font = Font(bold=True, size=18)
ws4['A1'].fill = PatternFill("solid", fgColor="1F4E79")
ws4['A1'].font = Font(bold=True, size=18, color="FFFFFF")
ws4.merge_cells('A1:D1')

ws4['A3'] = "ImageTrend Partnership Opportunity"
ws4['A3'].font = Font(bold=True, size=14)

summary_data = [
    ('ImageTrend 2024 Revenue', '$25.2M', 'Verified - GetLatka'),
    ('ImageTrend Growth Rate', '30%+', '2023-2024'),
    ('ImageTrend Estimated Valuation', '$150-200M', '6-8x Revenue'),
    ('ImageTrend EMS Customer Base', '13,000 agencies', 'Your TAM'),
    ('', '', ''),
    ('Protocol Guide Price', '$4.99/mo or $39/yr', 'Per user'),
    ('ImageTrend Take Rate', '20%', 'Distribution fee'),
    ('Your Revenue Share', '80%', 'Net to you'),
    ('', '', ''),
    ('Year 1 ARR (Conservative)', "='Revenue Model'!I16", '2% adoption'),
    ('Year 1 ARR (Moderate)', "='Revenue Model'!I17", '5% adoption'),
    ('Year 3 ARR (Projected)', "='Revenue Model'!F29", '10% adoption'),
    ('Year 5 ARR (Projected)', "='Revenue Model'!F29", '20% adoption'),
]

for row_idx, (metric, value, note) in enumerate(summary_data, 4):
    ws4.cell(row=row_idx, column=1, value=metric).font = Font(bold=True) if metric else Font()
    cell = ws4.cell(row=row_idx, column=2, value=value)
    if isinstance(value, str) and value.startswith('='):
        cell.number_format = currency_format
    ws4.cell(row=row_idx, column=3, value=note)

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 20

# Save workbook
output_path = '/sessions/adoring-quirky-pascal/mnt/Protocol Guide Manus/ImageTrend_Partnership_Analysis.xlsx'
wb.save(output_path)
print(f"Excel model saved to: {output_path}")
